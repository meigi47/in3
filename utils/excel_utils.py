#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from utils.constants.constants import Constants
from openpyxl import Workbook, load_workbook
import os

from utils.logger import Logger
from utils.str_utils import str_utils


class Excel_load:
    def __init__(self):
        pass

    def load_wk(self, wbPath: str, stIndex=0) -> list:  # xlsx ->list; 还需要 yaml -> xlsx; xlsx -> yaml; yaml -> list
        """
        :param wbPath: excel路径
        :type wbPath: str

        :param stIndex: sheet索引（从0开始）
        :type stIndex: int
        """
        wb = load_workbook(wbPath)
        if stIndex >= len(wb.sheetnames):
            Logger.error_exit("excel表sheet的索引超出最大值。请重新设置")  #TODO:优先级有空格、期望结果的字段名有问题和读取方式应当为函数
        # 指定去找哪个表
        sheet = wb.worksheets[int(stIndex)]
        max_row = sheet.max_row
        # 以下两个都是读取sheet的单个格子用的
        results = []

        title_cols_dict = {}
        # 拿到第一行的title放进dict
        for title_cell in sheet[1]:
            title_cols_dict[title_cell.col_idx] = title_cell.value

        temp_row_dict = {}
        # 从第二行开始遍历数据，匹配title组装数据的dict
        for row in sheet.iter_rows(min_row=2, max_row=max_row):
            for row_cell in row:
                title_name = title_cols_dict.get(row_cell.col_idx)
                # 匹配需要转json的列
                if title_name in Constants.CONVERT_JSON_FIELD_LIST:
                    temp_row_dict[title_name] = str_utils.json_str_to_list(row_cell.value)
                else:
                    temp_row_dict[title_name] = row_cell.value

            results.append(temp_row_dict)
            temp_row_dict = {}
        # print(results)
        return results

    

    def write_wk(self, lists: list, out_file_name: str):
        wb = Workbook()
        # 用active初始化一个表
        ws = wb.active
        # 在wb上创建一个表sheet，这是 分外创建的
        #wsl = wb.create_sheet('Test')

        # 给表重新命名
        ws.title = os.path.basename(out_file_name)
        # 修改表命名的颜色
        ws.sheet_properties.tabColor = '1072BA'

        DATA = self.init_list(lists)
        # print(DATA)
        for row in DATA:
            ws.append(row)
        # 保存输出文件
        wb.save(out_file_name + ('' if  out_file_name.endswith('.xlsx') else '.xlsx'))

    def init_list(self, lists: list) -> list:
        """
        进来的list的结构是这样的：
        [{"Jira号":123}，{"Jira号":456}]
        """
        results = []
        title_list = list(dict(lists[0]).keys())
        results.append(title_list)
        for item_list in lists:
            itme_dict = dict(item_list)
            itme_dict['数据'] = str_utils.json_list_to_str(itme_dict['数据'])
            itme_dict['步骤'] = str_utils.json_list_to_str(itme_dict['步骤'])
            itme_dict['期望结果'] = str_utils.json_list_to_str(item_list['期望结果'])
            itme_dict['期望校验数据'] = str_utils.json_list_to_str(item_list['期望校验数据'])
            results.append(list(itme_dict.values()))

        return results


excel_load = Excel_load()

if __name__ == "__main__":
    excel_load.load_wk()
