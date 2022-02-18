from requests.api import head
from interface.abstract_prod_op import Abstract_production_obj
import requests
import os
from openpyxl import load_workbook
import warnings
import csv
from utils.global_data import global_data
from manager.request_manager import request_manager
import utils.commons as commons
from interface.cofa.oss import oss
from utils.logger import Log, Logger
import uuid
from urllib3 import encode_multipart_formdata
 
class Task(Abstract_production_obj):
    def __init__(self) -> None:
        self.data = {}
        self.module = "tskt"
        self.operation = "task"
        
    def download_file(self,url,path):#该方法用于下载文件并存放在指定的本地目录中
        res = requests.get(url)
        with open(path, 'wb') as f:
            f.write(res.content)

    def check_export_task(self,url,filename,total_elements):#该方法用于校验导出任务是否功能正常：校验文件内数据量与期望数据量是否一致
        #判断是否在result存在download_files目录，没有就创建
        path="result/download_files"
        if os.path.isdir(path):
            pass
        else:
            os.mkdir("result/download_files")
        self.download_file(url,path+"/"+filename)#该方法用于下载文件并存放在指定的本地目录中
        if ".xlsx" in filename:#若文件后缀名为xlsx
            with warnings.catch_warnings(record=True):#屏蔽告警：若不加这个会报：with warnings.catch_warnings(record=True)，不屏蔽不影响运行通过
                wb=load_workbook(path+"/"+filename)
                sheets = wb.worksheets
                sheet1 = sheets[0]
                rows=sheet1.max_row
            assert total_elements==rows-1
        if ".csv" in filename:#若文件后缀名为csv
            total=0
            with open(path+"/"+filename, 'r',errors='ignore') as f:
                csv_reader = csv.reader(f)
                for row in csv_reader:
                    total += 1
                assert total_elements==total-1

    def create(self, data: dict) -> dict:
        tmp = global_data.join_url_type(
            data, self.module, self.operation, "create")
        return request_manager.do_request(tmp)   

    def import_and_create_task(self, file_path,task_code):
        #该方法用于将文件导入到系统并创建任务
        file_name=file_path.split("/")[-1]
        #获取callback参数信息
        r=oss.get_callback_parms({"file_name":file_name})
        assert (r.status_code == 200)
        callbackParams=r.json()["callbackParams"]
        #获取上传文件的签名目录
        app_id=30#In3系统服务id，写死
        key="BC_IMPORT_SOURCE"
        r=oss.get_post_policy({"app_id":app_id,"key":key})
        assert (r.status_code == 200)
        #包装数据
        post_data={}
        post_data["key"]=r.json()["dir"]+"/"+(uuid.uuid4().hex)[-32:]
        post_data["policy"]=r.json()["policy"]
        post_data["OSSAccessKeyId"]=r.json()["accessid"]
        post_data["success_action_status"]=200
        post_data["callback"]=callbackParams
        post_data["signature"]=r.json()["signature"]
        post_data["Content-Disposition"]="attachment;filename="+file_name
        f=open(file_path,'rb')
        file=f.read()
        post_data["file"]=(file_name,file,"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")#若上传的文件不是xlsx文件，则需要对content-type进行修改
        
        boundary='----WebKitFormBoundary' + (uuid.uuid4().hex)[-16:]
        
        #发生请求到阿里云oss服务器
        #url="https://industics-test.oss-cn-shanghai.aliyuncs.com"
        upload_url=r.json()["host"]
        Logger.info(upload_url)
        dd=encode_multipart_formdata(post_data,boundary)[0]
        header={"Content-Type":encode_multipart_formdata(post_data,boundary)[1]}
        r=requests.post(url=upload_url,data=dd,headers=header)
        Logger.info(r.text)
        assert (r.status_code == 200)
        #file.close()
        import_object_id = r.json()["id"]
        #创建任务
        f.close()
        task_data={}
        task_data["code"]=task_code
        task_data["import_object_id"]=import_object_id
        r=self.create(task_data)
        assert (r.status_code == 200)
        task_id=r.json()["id"]
        return task_id
        


task = Task()
