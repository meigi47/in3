from logging import fatal
from re import T
from typing import Sequence
from case_data.abstract_case_mgr import Abstrct_case_mgr
from manager.abstract_case import Abstract_case
from interface.sd.contract import contract
from interface.md.customer import customer
from interface.sd.sales_order import sales_order
from interface.uaas.uaas import account
import pytest
from utils.global_data import Global_data, global_data
from utils.logger import Log, Logger
from interface.md.mdm_material import material
from interface.pp.work_order import work_order
from interface.pp.demand_supply import demand_supply
from interface.pp.mrp_rules import mrp_rules
from interface.pe.area import area
from interface.pe.route import route
from interface.tskt.task import task
from interface.wm.mo import mo
from interface.wm.mo_warehouse_application import mo_warehouse_application
from openpyxl import Workbook,load_workbook
import time


class Manager(Abstrct_case_mgr):
    def __init__(self) -> None:
        super().__init__()

    class create_workorder_by_salesorder(Abstract_case): 
        def __init__(self, text=None) -> None:
            super().__init__('计划管理/工单管理/新建工单/关联销售订单', text)
        def run(self):
            #根据已创建的销售合同，创建销售订单（前置条件，需满足），获取销售订单id
            
            ordered_id = account.get_field_by_name(self.case_data['数据'][0],"id")
            r = contract.retrieve(self.case_data['数据'][1])
            if r.json()["total_elements"] == 0:
                Logger.error('未在系统准备必要的销售合同')
                pytest.fail()
            contract_id=r.json()["content"][0]["id"]
            contract_no=r.json()["content"][0]["contract_no"]
            contract_name=r.json()["content"][0]["project_name"]
            account_name=r.json()["content"][0]["account_name"]
            customer_id=customer.get_field_by_name(self.case_data["数据"][2],"id")           
            self.case_data["数据"][3]["orderedBy"] = ordered_id
            self.case_data["数据"][3]["contractId"] = contract_id
            self.case_data["数据"][3]["account_id"] = customer_id

            #关于物料的维护：选择每次自动创建，并去更新其BOM。            
            #根据计量单位组，查询主计量单位和采购单位，这两个单位都由unit_name得来
            r=material.material_unit_groups_page(self.case_data["数据"][5])
            self.response_info(r)
            material_unit_groups_id=r.json()['content'][0]['id']
            r=material.material_units_retrieve_by_measure_unit_groups({"id":material_unit_groups_id})
            self.response_info(r)
            if len(r.json())==0:
                Logger.error('未在系统准备必要的主计量单位')
                pytest.fail()
            material_units_name=r.json()[0]["unit"]["unit_name"]
            material_list=[]
            #循环创建所需要的物料
            for caseData in self.case_data['数据']:
                if "物料名称" in caseData:
                    caseData["material_unit"]=material_units_name
                    caseData["material_purchase_unit"]=material_units_name
                    r=material.create(caseData)
                    self.response_info(r)
                    material_info={}
                    material_info["material_id"]=r.json()["id"]
                    material_info["material_no"]=r.json()["material_no"]
                    material_list.append(material_info)
            #修改物料的BOM
            BOM_info=material.get_bom_from_casedata(self.case_data['数据'][4]["BOM"])#获取物料BOM的关系
            #根据BOM去维护系统中物料的BOM的关系
            if len(BOM_info)!=0:
                for bom_info in BOM_info:
                    material.update_bom(bom_info)
            #物料相关信息的封装
            #搜索物料，获取其中的参数值
            self.case_data["数据"][3]["products"]=[]
            sequence=1
            if len(self.case_data["数据"][6]["交付物"])==0:
                Logger.error('未在测试数据准备交付物的数据')
                pytest.fail()
            for material_search in self.case_data["数据"][6]["交付物"]:#处理销售订单带多个交付物的情况
                r=material.retrieve({"material_no":material_search["物料编号"]})
                product={}
                product["material_group"] = r.json()["content"][0]["material_group"]
                product["material_desc"] = r.json()["content"][0]["material_desc"]
                product["material_name"] = r.json()["content"][0]["material_name"]
                product["material_no"] = r.json()["content"][0]["material_no"]
                product["material_id"] = r.json()["content"][0]["id"]
                product["measure_unit"] = r.json()["content"][0]["purchase_measure_unit"]
                product["unit_price"]=material_search["未税单价"]
                product["tax_rate"]=material_search["税率"]
                product["unit_tax_price"]=material_search["含税单价"]
                product["total_tax_price"]=material_search["价税合计"]
                product["quantity"]=material_search["数量"]
                product["required_delivery_date"]=material_search["交货日期"]
                product["price_mode"]="UNIT_PRICE"
                product["material_unit_name"]=r.json()["content"][0]["material_unit"]
                product["total_price"]=material_search["未税总额"]
                product["sequence"]= sequence
                sequence=sequence+1
                self.case_data["数据"][3]["products"].append(product)
            r = sales_order.create(self.case_data["数据"][3])
            self.response_info(r,"创建销售订单")
            assert (len(r.json()['new_id']) > 0)
            so_id=r.json()['new_id']
            #下发该订单
            #先默认系统内销售订单审批流是关闭的
            r=sales_order.create_approval({"task_config":None,"back_to_pr_flag":None,"so_id":so_id})
            self.response_info(r)
            
            #新建工单
            #内容包含 合同信息，客户信息，物料信息，订单信息，子物料信息，其他信息
            for product_info in self.case_data["数据"][3]["products"]:
                material_id=product_info["material_id"]
                material_name=product_info["material_name"]
                material_no=product_info["material_no"]
                #获取区域，id和名字
                r=area.retrieve_areas({"material_no":material_no})
                self.response_info(r)
                area_id=r.json()[0]["id"]
                area_name=r.json()[0]["area_name"]
                #获取工艺路线，id和名字
                r=route.retrieve_routes({"material_no":material_no})
                self.response_info(r)
                route_id=r.json()[0]["id"]
                route_name=r.json()[0]["route_name"]
                
                
                #获取主数据物料BOM树
                r=work_order.retrieve_bomm_tree([material_id])
                self.response_info(r)
                if "children" in r.json()[0]:#存在所添加的主数据没有子物料的情况，此处进行了处理。
                    children_list=r.json()[0]["children"]
                    material.recursiom_mixed_bom_data(children_list,1,0,None,[])
                    kitting_status=0
                    '''
                    n=0
                    kitting_status=0
                    for child in children_list:#此处对子物料的信息进行包装处理
                        i=n
                        child["id"]=child["id"]+"_"+ str(i)
                        del child["parent_material_id"]
                        #del child["procurement_type"]
                        if "procurement_type" in child:
                            del child["procurement_type"]  
                        child["kitting_status"]=0
                        child["sequence"]=n
                        child["level"]=1
                        child["expanded"]=True
                        child["location"]=True
                        row=20+n
                        child["xid"]="row_"+ str(row)
                        n=n+1
                    '''
                else:
                    children_list=[]
                    kitting_status=10
                Logger.info(children_list)

                
                #获取订单信息，交付物数量，id，no，名字，交期
                r=sales_order.get_detail({'so_id': so_id})
                batch_delivery_date=r.json()["batch_delivery_date"]
                so_no=r.json()["batch_no"]
                so_name=r.json()["batch_name"]
                #不同的交付物信息不一样，需要处理
                document_item_id=""
                total_quantity=0
                for prod in r.json()["products"]:
                    if prod["material_id"]==material_id:
                        document_item_id=prod["id"]+document_item_id
                        total_quantity=prod["quantity"]+total_quantity
                #包装请求
                data={}
                data["reference_document_type"]="RELATED_SO"
                data["contract_name"]=contract_name
                data["contract_no"]=contract_no
                data["account_name"]=account_name
                data["material_id"]=material_id
                data["material_name"]=material_name
                data["material"]={
                    "id":material_id,
                    "material_name":material_name,
                    "material_no":material_no
                }
                data["mo_type"]="PRODUCTION_ORDER"
                data["pe_route_id"]=route_id
                data["pe_area_id"]=area_id
                data["plan_prod_date"]="2021-11-29T00:00:00"
                data["plan_stock_date"]="2021-11-30T00:00:00"
                data["document_date"]=batch_delivery_date
                data["document_type"]="SO"
                data["document_id"]=so_id
                data["document_no"]=so_no
                data["document_name"]=so_name
                data["document_item_id"]=document_item_id
                data["pe_route_name"]=route_name
                data["pe_area_name"]=area_name
                data["completed_quantity"]=0
                data["kitting_status"]=kitting_status
                data["kitting_items"]=children_list
                data["total_quantity"]=total_quantity
                r=work_order.create(data)
                self.response_info(r)
                #校验
                #需要校验的数据有：销售订单名称，参考单据类型，客户名称，订单交期，合同编号，产品编号，生产数量，计划完成日期，工单bom关系，物料需求明细
                work_order_id=r.json()["id"]
                r=work_order.get_detail({"production_plan_id":work_order_id})
                self.response_info(r)

                #关于基本校验信息的包装
                check_info={}
                check_info["销售订单名称"]=r.json()["document_name"]
                check_info["参考单据类型"]=r.json()["reference_document_type_desc"]
                check_info["客户名称"]=r.json()["account_name"]
                check_info["订单交期"]=r.json()["document_date"]
                check_info["合同编号"]=r.json()["contract_no"]
                check_info["生产数量"]=r.json()["total_quantity"]
                check_info["计划完成日期"]=r.json()["plan_stock_date"]

                #关于BOM校验信息的包装
                check_bom=material.get_bom_from_wo_detail(r.json())#通过该方法获取查询到的工单详情的bom信息，返回一个如[{'A': [{'B': 1}, {'C': 1}]}]的列表去表示bom

                #关于物料明细校验信息的包装：直接由测试数据得出应有的物料明细即可,给出内容为：物料编号，工单需求，单台需求
                check_material_info=[]
                if len(r.json()["bom_qty_list"])!=0:
                    for bom_qty in r.json()["bom_qty_list"]:
                        tmp={}
                        tmp["物料编号"]=bom_qty["material_no"]
                        tmp["工单需求"]=bom_qty["required_qty"]
                        tmp["单台需求"]=bom_qty["required_unit_qty"]
                        check_material_info.append(tmp)

                bom_compare=0
                is_check_data_bom=0
                for check_data in self.case_data['期望校验数据']:
                    if check_data["交付物"]==material_no:
                        self.compare_result_detail(check_info,check_data)#校验基本信息
                        BOM_info=material.get_bom_from_casedata(check_data["BOM"])#获取物料BOM的关系
                        bom_compare=self.is_list_same(check_bom,BOM_info)#比较物料BOM是否一致
                        if bom_compare==0:
                            Logger.info("工单BOM与期望数据不一致")
                            pytest.fail()
                        #比较物料明细信息是否一致
                        if len(check_data["物料明细"])==0 and len(check_material_info)==0:#没有给出期望的物料明细信息以及工单详情没有物料明细信息
                            pass
                        elif len(check_data["物料明细"])==0:
                            Logger.info("期望数据未给出产出物"+material_no+"的物料明细期望校验值")
                            pytest.fail()
                        elif len(check_material_info)==0:
                            Logger.info(material_no+"给出物料明细期望校验值，但工单内未有物料明细值")
                            pytest.fail()
                        elif self.is_list_same(check_material_info,check_data["物料明细"])==0:
                            Logger.info(material_no+"的物料明细期望校验值与实际不一致")
                            pytest.fail()
                        is_check_data_bom=1
                if is_check_data_bom==0:
                    Logger.info("期望数据未给出产出物"+material_no+"的期望校验值")
                    pytest.fail()
                
            
            
    class create_workorder_bind_salesorder(Abstract_case): 
        def __init__(self, text=None) -> None:
            super().__init__('计划管理/工单管理/新建工单/手绑销售订单', text)
        def run(self):
            #根据已创建的销售合同，创建销售订单（前置条件，需满足），获取销售订单id
            ordered_id = account.get_field_by_name(self.case_data['数据'][0],"id")
            r = contract.retrieve(self.case_data['数据'][1])
            if r.json()["total_elements"] == 0:
                Logger.error('未在系统准备必要的销售合同')
                pytest.fail()
            contract_id=r.json()["content"][0]["id"]
            contract_no=r.json()["content"][0]["contract_no"]
            contract_name=r.json()["content"][0]["project_name"]
            account_name=r.json()["content"][0]["account_name"]
            customer_id=customer.get_field_by_name(self.case_data["数据"][2],"id")           
            self.case_data["数据"][3]["orderedBy"] = ordered_id
            self.case_data["数据"][3]["contractId"] = contract_id
            self.case_data["数据"][3]["account_id"] = customer_id

            #关于物料的维护：选择每次自动创建，并去更新其BOM。            
            #根据计量单位组，查询主计量单位和采购单位，这两个单位都由unit_name得来
            r=material.material_unit_groups_page(self.case_data["数据"][5])
            self.response_info(r)
            material_unit_groups_id=r.json()['content'][0]['id']
            r=material.material_units_retrieve_by_measure_unit_groups({"id":material_unit_groups_id})
            self.response_info(r)
            if len(r.json())==0:
                Logger.error('未在系统准备必要的主计量单位')
                pytest.fail()
            material_units_name=r.json()[0]["unit"]["unit_name"]
            material_list=[]
            #循环创建所需要的物料
            for caseData in self.case_data['数据']:
                if "物料名称" in caseData:
                    caseData["material_unit"]=material_units_name
                    caseData["material_purchase_unit"]=material_units_name
                    r=material.create(caseData)
                    self.response_info(r)
                    material_info={}
                    material_info["material_id"]=r.json()["id"]
                    material_info["material_no"]=r.json()["material_no"]
                    material_list.append(material_info)
            #修改物料的BOM
            BOM_info=material.get_bom_from_casedata(self.case_data['数据'][4]["BOM"])#获取物料BOM的关系
            #根据BOM去维护系统中物料的BOM的关系
            if len(BOM_info)!=0:
                for bom_info in BOM_info:
                    material.update_bom(bom_info)
            #物料相关信息的封装
            #搜索物料，获取其中的参数值
            self.case_data["数据"][3]["products"]=[]
            sequence=1
            if len(self.case_data["数据"][6]["交付物"])==0:
                Logger.error('未在测试数据准备交付物的数据')
                pytest.fail()
            for material_search in self.case_data["数据"][6]["交付物"]:#处理销售订单带多个交付物的情况
                r=material.retrieve({"material_no":material_search["物料编号"]})
                product={}
                product["material_group"] = r.json()["content"][0]["material_group"]
                product["material_desc"] = r.json()["content"][0]["material_desc"]
                product["material_name"] = r.json()["content"][0]["material_name"]
                product["material_no"] = r.json()["content"][0]["material_no"]
                product["material_id"] = r.json()["content"][0]["id"]
                product["measure_unit"] = r.json()["content"][0]["purchase_measure_unit"]
                product["unit_price"]=material_search["未税单价"]
                product["tax_rate"]=material_search["税率"]
                product["unit_tax_price"]=material_search["含税单价"]
                product["total_tax_price"]=material_search["价税合计"]
                product["quantity"]=material_search["数量"]
                product["required_delivery_date"]=material_search["交货日期"]
                product["price_mode"]="UNIT_PRICE"
                product["material_unit_name"]=r.json()["content"][0]["material_unit"]
                product["total_price"]=material_search["未税总额"]
                product["sequence"]= sequence
                sequence=sequence+1
                self.case_data["数据"][3]["products"].append(product)
            r = sales_order.create(self.case_data["数据"][3])
            self.response_info(r)
            assert (len(r.json()['new_id']) > 0)
            so_id=r.json()['new_id']
            #下发该订单
            #先默认系统内销售订单审批流是关闭的
            r=sales_order.create_approval({"task_config":None,"back_to_pr_flag":None,"so_id":so_id})
            self.response_info(r)
            #新建工单
            #内容包含 合同信息，客户信息，物料信息，订单信息，子物料信息，其他信息
            for product_info in self.case_data["数据"][3]["products"]:
                material_id=product_info["material_id"]
                material_name=product_info["material_name"]
                material_no=product_info["material_no"]
                #获取区域，id和名字
                r=area.retrieve_areas({"material_no":material_no})
                self.response_info(r)
                area_id=r.json()[0]["id"]
                area_name=r.json()[0]["area_name"]
                #获取工艺路线，id和名字
                r=route.retrieve_routes({"material_no":material_no})
                self.response_info(r)
                route_id=r.json()[0]["id"]
                route_name=r.json()[0]["route_name"]
                #获取主数据物料BOM树
                r=work_order.retrieve_bomm_tree([material_id])
                self.response_info(r)
                if "children" in r.json()[0]:#存在所添加的主数据没有子物料的情况，此处进行了处理。
                    children_list=r.json()[0]["children"]
                    material.recursiom_mixed_bom_data(children_list,1,0,None,[])
                    kitting_status=0
                    '''
                    for child in children_list:#此处对子物料的信息进行包装处理
                        i=n
                        child["id"]=child["id"]+"_"+ str(i)
                        del child["parent_material_id"]
                        #del child["procurement_type"]
                        if "procurement_type" in child:
                            del child["procurement_type"]  
                        child["kitting_status"]=0
                        child["sequence"]=n
                        child["level"]=1
                        child["expanded"]=True
                        child["location"]=True
                        row=20+n
                        child["xid"]="row_"+ str(row)
                        n=n+1
                    '''
                else:
                    children_list=[]
                    kitting_status=10
                Logger.info(children_list)
                
                #获取订单信息，交付物数量，id，no，名字，交期
                r=sales_order.get_detail({'so_id': so_id})
                batch_delivery_date=r.json()["batch_delivery_date"]
                so_no=r.json()["batch_no"]
                so_name=r.json()["batch_name"]
                #不同的交付物信息不一样，需要处理
                document_item_id=""
                total_quantity=0
                for prod in r.json()["products"]:
                    if prod["material_id"]==material_id:
                        document_item_id=prod["id"]+document_item_id
                        total_quantity=prod["quantity"]+total_quantity
                #包装请求
                data={}
                data["reference_document_type"]="HAND_TIED_SO"
                data["contract_name"]=contract_name
                data["contract_no"]=contract_no
                data["account_name"]=account_name
                data["material_id"]=material_id
                data["material_name"]=material_name
                data["material"]={
                    "id":material_id,
                    "material_name":material_name,
                    "material_no":material_no
                }
                data["mo_type"]="PRODUCTION_ORDER"
                data["pe_route_id"]=route_id
                data["pe_area_id"]=area_id
                data["plan_prod_date"]="2021-11-29T00:00:00"
                data["plan_stock_date"]="2021-11-30T00:00:00"
                data["document_date"]=batch_delivery_date
                data["document_type"]="SO"
                data["document_id"]=so_id
                data["document_no"]=so_no
                data["document_name"]=so_name
                data["document_item_id"]=document_item_id
                data["pe_route_name"]=route_name
                data["pe_area_name"]=area_name
                data["completed_quantity"]=0
                data["kitting_status"]=kitting_status
                data["kitting_items"]=children_list
                data["total_quantity"]=total_quantity
                r=work_order.create(data)
                self.response_info(r)
                #校验
                #需要校验的数据有：销售订单名称，参考单据类型，客户名称，订单交期，合同编号，产品编号，生产数量，计划完成日期，工单bom关系，物料需求明细
                work_order_id=r.json()["id"]
                r=work_order.get_detail({"production_plan_id":work_order_id})
                self.response_info(r)

                #关于基本校验信息的包装
                check_info={}
                check_info["销售订单名称"]=r.json()["document_name"]
                check_info["参考单据类型"]=r.json()["reference_document_type_desc"]
                check_info["客户名称"]=r.json()["account_name"]
                check_info["订单交期"]=r.json()["document_date"]
                check_info["合同编号"]=r.json()["contract_no"]
                check_info["生产数量"]=r.json()["total_quantity"]
                check_info["计划完成日期"]=r.json()["plan_stock_date"]

                #关于BOM校验信息的包装
                check_bom=material.get_bom_from_wo_detail(r.json())#通过该方法获取查询到的工单详情的bom信息，返回一个如[{'A': [{'B': 1}, {'C': 1}]}]的列表去表示bom

                #关于物料明细校验信息的包装：直接由测试数据得出应有的物料明细即可,给出内容为：物料编号，工单需求，单台需求
                check_material_info=[]
                if len(r.json()["bom_qty_list"])!=0:
                    for bom_qty in r.json()["bom_qty_list"]:
                        tmp={}
                        tmp["物料编号"]=bom_qty["material_no"]
                        tmp["工单需求"]=bom_qty["required_qty"]
                        tmp["单台需求"]=bom_qty["required_unit_qty"]
                        check_material_info.append(tmp)

                bom_compare=0
                is_check_data_bom=0
                for check_data in self.case_data['期望校验数据']:
                    if check_data["交付物"]==material_no:
                        self.compare_result_detail(check_info,check_data)#校验基本信息
                        BOM_info=material.get_bom_from_casedata(check_data["BOM"])#获取物料BOM的关系
                        bom_compare=self.is_list_same(check_bom,BOM_info)#比较物料BOM是否一致
                        if bom_compare==0:
                            Logger.info("工单BOM与期望数据不一致")
                            pytest.fail()
                        #比较物料明细信息是否一致
                        if len(check_data["物料明细"])==0 and len(check_material_info)==0:#没有给出期望的物料明细信息以及工单详情没有物料明细信息
                            pass
                        elif len(check_data["物料明细"])==0:
                            Logger.info("期望数据未给出产出物"+material_no+"的物料明细期望校验值")
                            pytest.fail()
                        elif len(check_material_info)==0:
                            Logger.info(material_no+"给出物料明细期望校验值，但工单内未有物料明细值")
                            pytest.fail()
                        elif self.is_list_same(check_material_info,check_data["物料明细"])==0:
                            Logger.info(material_no+"的物料明细期望校验值与实际不一致")
                            pytest.fail()
                        is_check_data_bom=1
                if is_check_data_bom==0:
                    Logger.info("期望数据未给出产出物"+material_no+"的期望校验值")
                    pytest.fail()
   

    class retrieve_workorder(Abstract_case): 
        def __init__(self, text=None) -> None:
            super().__init__('计划管理/工单管理/查询工单', text)
        def run(self):
            #查询工单列表
            #根据测试数据（销售订单名称+物料名称）进行查询工单，返回的数据与校验数据（销售订单名称+物料id）进行对比
            #查询销售订单名称查询，获取销售订单的信息
            r=sales_order.retrieve(self.case_data['数据'][0])
            self.response_info(r)
            if r.json()["total_elements"] == 0:
                Logger.error('查询的订单未创建')
                pytest.fail()
            #此处存在查询会返回多个值的问题，获取返回值中与查询数据完全一致的数据
            so_info=self.query_directly(r.json()["content"],"batch_name",self.case_data['数据'][0]["销售订单名称"])
            data={}
            data["so_id"]=so_info["id"]
            #查询物料信息
            r=material.retrieve(self.case_data["数据"][1])
            self.response_info(r)
            if r.json()["total_elements"] == 0:
                Logger.error('未在系统准备必要的物料')
                pytest.fail()
            #此处存在查询会返回多个值的问题，获取返回值中与查询数据完全一致的数据
            material_info=self.query_directly(r.json()["content"],"material_name",self.case_data['数据'][1]["物料名称"])
            data["material_ids"]=[material_info["id"]]
            #下发搜索信息
            data["kitting_status"]=0
            data["shipped_status"]=0
            data["finish_status"]=0
            data["need_update_bom"]=False
            data["production_plan_status"]=[0,10]
            data["size"]=20
            data["page"]=0
            r=work_order.retrieve(data)
            self.response_info(r)
            #与校验数据进行校验
            self.compare_result_page(self.case_data['期望校验数据'][0],r)
            self.compare_result_page({"material_id":material_info["id"]},r)#根据之前查询获得的物料id进行校验

    class detail_of_workorder(Abstract_case): 
        def __init__(self, text=None) -> None:
            super().__init__('计划管理/工单管理/工单详情', text)
        def run(self):
            #查看工单详情
            #根据测试数据（销售订单名称+物料名称）进行查询工单，返回的数据与校验数据（销售订单名称+物料id）进行对比
            #查询销售订单名称查询，获取销售订单的信息
            r=sales_order.retrieve(self.case_data['数据'][0])
            self.response_info(r)
            if r.json()["total_elements"] == 0:
                Logger.error('查询的订单未创建')
                pytest.fail()
            #此处存在查询会返回多个值的问题，获取返回值中与查询数据完全一致的数据
            so_info=self.query_directly(r.json()["content"],"batch_name",self.case_data['数据'][0]["销售订单名称"])
            data={}
            data["so_id"]=so_info["id"]
            #查询物料信息
            r=material.retrieve(self.case_data["数据"][1])
            self.response_info(r)
            if r.json()["total_elements"] == 0:
                Logger.error('未在系统准备必要的物料')
                pytest.fail()
            #此处存在查询会返回多个值的问题，获取返回值中与查询数据完全一致的数据
            material_info=self.query_directly(r.json()["content"],"material_name",self.case_data['数据'][1]["物料名称"])
            data["material_ids"]=[material_info["id"]]
            #下发搜索信息
            data["kitting_status"]=0
            data["shipped_status"]=0
            data["finish_status"]=0
            data["need_update_bom"]=False
            data["production_plan_status"]=[0,10]
            data["size"]=20
            data["page"]=0
            r=work_order.retrieve(data)
            self.response_info(r)
            #获取工单id，从而获得工单详情
            work_order_info=self.query_directly(r.json()['content'],'document_name',self.case_data['数据'][0]["销售订单名称"])
            work_order_id=work_order_info["id"]
            r=work_order.get_detail({"production_plan_id":work_order_id})
            self.response_info(r)
            #与校验数据进行校验
            self.compare_result_detail(self.case_data['期望校验数据'][0],r.json())


    class create_workorder_import(Abstract_case): 
        def __init__(self, text=None) -> None:
            super().__init__('计划管理/工单管理/新建工单/excel导入创建工单', text)
        def run(self):
            #根据已创建的销售合同，创建销售订单（前置条件，需满足），获取销售订单id
            
        
            
            ordered_id = account.get_field_by_name(self.case_data['数据'][0],"id")
            r = contract.retrieve(self.case_data['数据'][1])
            if r.json()["total_elements"] == 0:
                Logger.error('未在系统准备必要的销售合同')
                pytest.fail()
            contract_id=r.json()["content"][0]["id"]
            contract_no=r.json()["content"][0]["contract_no"]
            contract_name=r.json()["content"][0]["project_name"]
            account_name=r.json()["content"][0]["account_name"]
            customer_id=customer.get_field_by_name(self.case_data["数据"][2],"id")           
            self.case_data["数据"][3]["orderedBy"] = ordered_id
            self.case_data["数据"][3]["contractId"] = contract_id
            self.case_data["数据"][3]["account_id"] = customer_id

            #关于物料的维护：选择每次自动创建，并去更新其BOM。            
            #根据计量单位组，查询主计量单位和采购单位，这两个单位都由unit_name得来
            r=material.material_unit_groups_page(self.case_data["数据"][5])
            self.response_info(r)
            material_unit_groups_id=r.json()['content'][0]['id']
            r=material.material_units_retrieve_by_measure_unit_groups({"id":material_unit_groups_id})
            self.response_info(r)
            if len(r.json())==0:
                Logger.error('未在系统准备必要的主计量单位')
                pytest.fail()
            material_units_name=r.json()[0]["unit"]["unit_name"]
            material_list=[]
            #循环创建所需要的物料
            for caseData in self.case_data['数据']:
                if "物料名称" in caseData:
                    caseData["material_unit"]=material_units_name
                    caseData["material_purchase_unit"]=material_units_name
                    r=material.create(caseData)
                    self.response_info(r)
                    material_info={}
                    material_info["material_id"]=r.json()["id"]
                    material_info["material_no"]=r.json()["material_no"]
                    material_list.append(material_info)
            #修改物料的BOM
            BOM_info=material.get_bom_from_casedata(self.case_data['数据'][4]["BOM"])#获取物料BOM的关系
            #根据BOM去维护系统中物料的BOM的关系
            if len(BOM_info)!=0:
                for bom_info in BOM_info:
                    material.update_bom(bom_info)
            #物料相关信息的封装
            #搜索物料，获取其中的参数值
            self.case_data["数据"][3]["products"]=[]
            sequence=1
            if len(self.case_data["数据"][6]["交付物"])==0:
                Logger.error('未在测试数据准备交付物的数据')
                pytest.fail()
            for material_search in self.case_data["数据"][6]["交付物"]:#处理销售订单带多个交付物的情况
                r=material.retrieve({"material_no":material_search["物料编号"]})
                product={}
                product["material_group"] = r.json()["content"][0]["material_group"]
                product["material_desc"] = r.json()["content"][0]["material_desc"]
                product["material_name"] = r.json()["content"][0]["material_name"]
                product["material_no"] = r.json()["content"][0]["material_no"]
                product["material_id"] = r.json()["content"][0]["id"]
                product["measure_unit"] = r.json()["content"][0]["purchase_measure_unit"]
                product["unit_price"]=material_search["未税单价"]
                product["tax_rate"]=material_search["税率"]
                product["unit_tax_price"]=material_search["含税单价"]
                product["total_tax_price"]=material_search["价税合计"]
                product["quantity"]=material_search["数量"]
                product["required_delivery_date"]=material_search["交货日期"]
                product["price_mode"]="UNIT_PRICE"
                product["material_unit_name"]=r.json()["content"][0]["material_unit"]
                product["total_price"]=material_search["未税总额"]
                product["sequence"]= sequence
                sequence=sequence+1
                self.case_data["数据"][3]["products"].append(product)
            r = sales_order.create(self.case_data["数据"][3])
            self.response_info(r)
            assert (len(r.json()['new_id']) > 0)
            so_id=r.json()['new_id']
            #下发该订单
            #先默认系统内销售订单审批流是关闭的
            r=sales_order.create_approval({"task_config":None,"back_to_pr_flag":None,"so_id":so_id})
            self.response_info(r)
            
            
            #在导入模板中写入数据
            #下载模板到case_data，若该路径已有文件，删除
            work_order.get_template_of_createWO("case_data/批量导入开工单模板_V1.0.xlsx")
            #写入信息
            wo_info_list=[]
            for product_info in self.case_data["数据"][3]["products"]:

                r=sales_order.get_detail({'so_id': so_id})
                so_no=r.json()["batch_no"]
                #获取区域，id和名字
                r=area.retrieve_areas({"material_no":product_info["material_no"]})
                self.response_info(r)
                area_id=r.json()[0]["id"]
                area_name=r.json()[0]["area_name"]
                #获取工艺路线，id和名字
                r=route.retrieve_routes({"material_no":product_info["material_no"]})
                self.response_info(r)
                route_id=r.json()[0]["id"]
                route_name=r.json()[0]["route_name"]
                data={}
                #data["参考单据类型"]="关联销售订单"
                data["销售订单编号"]=so_no
                data["*物料编号"]=product_info["material_no"]
                #data["物料名称"]=product_info["material_name"]
                #data["工单类型"]="生产工单"
                #data["区域"]=area_name
                #data["工艺路线"]=route_name
                data["*生产数量"]="100"
                data["*计划上线时间"]="2021/11/30"
                data["*计划完成时间"]="2021/12/30"  
                wo_info_list.append(data)         
            wb=load_workbook("case_data/批量导入开工单模板_V1.0.xlsx")
            ws=wb['导入表_批量开工单模板']
                  
            for wo in wo_info_list:
                w_data=[]
                for i in ws[1]:
                    if i.value in wo:
                        w_data.append(wo[i.value])
                    else:
                        w_data.append(None)
                ws.append(w_data)
                Logger.info(w_data)
            wb.save("case_data/批量导入开工单模板_V1.0.xlsx")
            
            
            #导入文件，创建任务
            task_id=task.import_and_create_task("case_data/批量导入开工单模板_V1.0.xlsx","PRODUCTION_PLAN_BATCH_IMPORT")
            
            #任务成功->获取导出文件地址exportFileUrl->校验
            status="CREATED"
            
            while status == "CREATED":
                time.sleep(5)#考虑到可能查询时候任务未完成，设置5秒一查
                r=task.retrieve({"size": 20,"page": 0})
                for items in r.json()["content"]:
                    if items["id"]==task_id:
                        if items["status"]=="CREATED":
                            break
                        elif items["status"]=="EXECUTING":
                            break
                        elif items["status"]=="SUCCEEDED":
                            status="SUCCEEDED"
                              
                            break
                        else:
                            Logger.info("批量工单创建失败")
                            pytest.fail()
    
    class update_bom_of_workorder_import(Abstract_case): 
        def __init__(self, text=None) -> None:
            super().__init__('计划管理/工单管理/批量更新工单BOM', text)
        def run(self):
            #读取测试数据，获取所需修改的bom内容：工单号，位号，父物料编号，子物料编号，数量。将数据包装，写入下载的物料模板内，导入。校验任务是否成功
            #根据测试数据（销售订单名称+物料名称）进行查询工单，返回的数据与校验数据（销售订单名称+物料id）进行对比
            #查询销售订单名称查询，获取销售订单的信息
            r=sales_order.retrieve(self.case_data['数据'][1])
            self.response_info(r)
            if r.json()["total_elements"] == 0:
                Logger.error('查询的订单未创建')
                pytest.fail()
            #此处存在查询会返回多个值的问题，获取返回值中与查询数据完全一致的数据
            so_info=self.query_directly(r.json()["content"],"batch_name",self.case_data['数据'][1]["销售订单名称"])
            data={}
            data["so_id"]=so_info["id"]
            #查询物料信息
            r=material.retrieve(self.case_data["数据"][0])
            self.response_info(r)
            if r.json()["total_elements"] == 0:
                Logger.error('未在系统准备必要的物料')
                pytest.fail()
            #此处存在查询会返回多个值的问题，获取返回值中与查询数据完全一致的数据
            material_info=self.query_directly(r.json()["content"],"material_name",self.case_data['数据'][0]["物料名称"])
            data["material_ids"]=[material_info["id"]]
            #下发搜索信息
            data["kitting_status"]=0
            data["shipped_status"]=0
            data["finish_status"]=0
            data["need_update_bom"]=False
            data["production_plan_status"]=[0,10]
            data["size"]=20
            data["page"]=0
            time.sleep(30)
            r=work_order.retrieve(data)
            self.response_info(r)
            work_order_info=self.query_directly(r.json()['content'],'document_name',self.case_data['数据'][1]["销售订单名称"])
            work_order_no=work_order_info["sn"]
            #包装信息写入到excel文件
            work_order.get_template_of_updateBOM("case_data/批量导入模板_工单BOM_V1.0.xlsx")
            wo_info_list=[]
            for bom_info in self.case_data["数据"][2]["BOM"]:
                data={}
                #data["参考单据类型"]="关联销售订单"
                
                data["*工单号"]=work_order_no
                data["*位号"]=bom_info["location_no"]
                data["*父件物料编号"]=bom_info["parent"]
                data["*子件物料编号"]=bom_info["children"]
                data["*单位用量"]=bom_info["number"]

                wo_info_list.append(data)         
            wb=load_workbook("case_data/批量导入模板_工单BOM_V1.0.xlsx")
            ws=wb['导入表_批量创建模板']
                  
            for wo in wo_info_list:
                w_data=[]
                for i in ws[1]:
                    if i.value in wo:
                        w_data.append(wo[i.value])
                    else:
                        w_data.append(None)
                ws.append(w_data)
                Logger.info(w_data)
            wb.save("case_data/批量导入模板_工单BOM_V1.0.xlsx")
            task_id=task.import_and_create_task("case_data/批量导入模板_工单BOM_V1.0.xlsx","MO_BOM_IMPORT")
            #任务成功->获取导出文件地址exportFileUrl->校验
            status="CREATED"
            
            while status == "CREATED":
                time.sleep(5)#考虑到可能查询时候任务未完成，设置5秒一查
                r=task.retrieve({"size": 20,"page": 0})
                for items in r.json()["content"]:
                    if items["id"]==task_id:
                        if items["status"]=="CREATED":
                            break
                        elif items["status"]=="EXECUTING":
                            break
                        elif items["status"]=="SUCCEEDED":
                            status="SUCCEEDED"
                              
                            break
                        else:
                            Logger.info("批量更新工单BOM失败")
                            pytest.fail()
            #校验

    class create_workorder_exchange(Abstract_case): 
        def __init__(self, text=None) -> None:
            super().__init__('计划管理/工单管理/新建工单/生产需求计划转工单', text)
        def run(self):
            #该方法实现的是单个转换，暂未做批量转换
            #该方法仅针对未设置“自动转MO"的情况。如设置需要取消
            #根据已创建的销售合同，创建销售订单（前置条件，需满足），获取销售订单id
            ordered_id = account.get_field_by_name(self.case_data['数据'][0],"id")
            r = contract.retrieve(self.case_data['数据'][1])
            self.response_info(r)
            if r.json()["total_elements"] == 0:
                Logger.error('未在系统准备必要的销售合同')
                pytest.fail()
            contract_id=r.json()["content"][0]["id"]
            contract_no=r.json()["content"][0]["contract_no"]
            contract_name=r.json()["content"][0]["project_name"]
            account_name=r.json()["content"][0]["account_name"]
            customer_id=customer.get_field_by_name(self.case_data["数据"][2],"id")           
            self.case_data["数据"][3]["orderedBy"] = ordered_id
            self.case_data["数据"][3]["contractId"] = contract_id
            self.case_data["数据"][3]["account_id"] = customer_id

            #关于物料的维护：选择每次自动创建，并去更新其BOM。            
            #根据计量单位组，查询主计量单位和采购单位，这两个单位都由unit_name得来
            r=material.material_unit_groups_page(self.case_data["数据"][5])
            self.response_info(r)
            material_unit_groups_id=r.json()['content'][0]['id']
            r=material.material_units_retrieve_by_measure_unit_groups({"id":material_unit_groups_id})
            self.response_info(r)
            if len(r.json())==0:
                Logger.error('未在系统准备必要的主计量单位')
                pytest.fail()
            material_units_name=r.json()[0]["unit"]["unit_name"]
            material_list=[]
            #循环创建所需要的物料
            for caseData in self.case_data['数据']:
                if "物料名称" in caseData:
                    caseData["material_unit"]=material_units_name
                    caseData["material_purchase_unit"]=material_units_name
                    r=material.create(caseData)
                    self.response_info(r)
                    material_info={}
                    material_info["material_id"]=r.json()["id"]
                    material_info["material_no"]=r.json()["material_no"]
                    material_list.append(material_info)
            #修改物料的BOM
            BOM_info=material.get_bom_from_casedata(self.case_data['数据'][4]["BOM"])#获取物料BOM的关系
            #根据BOM去维护系统中物料的BOM的关系
            if len(BOM_info)!=0:
                for bom_info in BOM_info:
                    material.update_bom(bom_info)
            #物料相关信息的封装
            #搜索物料，获取其中的参数值
            self.case_data["数据"][3]["products"]=[]
            sequence=1
            if len(self.case_data["数据"][6]["交付物"])==0:
                Logger.error('未在测试数据准备交付物的数据')
                pytest.fail()
            for material_search in self.case_data["数据"][6]["交付物"]:#处理销售订单带多个交付物的情况
                r=material.retrieve({"material_no":material_search["物料编号"]})
                product={}
                product["material_group"] = r.json()["content"][0]["material_group"]
                product["material_desc"] = r.json()["content"][0]["material_desc"]
                product["material_name"] = r.json()["content"][0]["material_name"]
                product["material_no"] = r.json()["content"][0]["material_no"]
                product["material_id"] = r.json()["content"][0]["id"]
                product["measure_unit"] = r.json()["content"][0]["purchase_measure_unit"]
                product["unit_price"]=material_search["未税单价"]
                product["tax_rate"]=material_search["税率"]
                product["unit_tax_price"]=material_search["含税单价"]
                product["total_tax_price"]=material_search["价税合计"]
                product["quantity"]=material_search["数量"]
                product["required_delivery_date"]=material_search["交货日期"]
                product["price_mode"]="UNIT_PRICE"
                product["material_unit_name"]=r.json()["content"][0]["material_unit"]
                product["total_price"]=material_search["未税总额"]
                product["sequence"]= sequence
                sequence=sequence+1
                self.case_data["数据"][3]["products"].append(product)
            r = sales_order.create(self.case_data["数据"][3])
            self.response_info(r)
            assert (len(r.json()['new_id']) > 0)
            so_id=r.json()['new_id']
            #下发该订单
            #先默认系统内销售订单审批流是关闭的
            r=sales_order.create_approval({"task_config":None,"back_to_pr_flag":None,"so_id":so_id})
            self.response_info(r)
            #跑MRP,判断MRP计算是否成功，若成功继续，失败则报错
            r=task.create({"code":"MRP_COMPUTE"})
            self.response_info(r)
            task_id=r.json()["id"]
            r=task.retrieve({"size": 20,"page": 0})
            self.response_info(r)
            assert r.json()["total_elements"]!=0
            status="CREATED"
            while status == "CREATED":
                time.sleep(5)#考虑到可能查询时候任务未完成，设置5秒一查
                r=task.retrieve({"size": 20,"page": 0})
                for items in r.json()["content"]:
                    if items["id"]==task_id:
                        if items["status"]=="CREATED":
                            break
                        elif items["status"]=="EXECUTING":
                            break
                        elif items["status"]=="SUCCEEDED":
                            status="SUCCEEDED" 
                            break
                        else:
                            Logger.info("MRP计算失败")
                            pytest.fail()
            #此时MRP计算完成，应已转化为生产需求计划

            #做是否已自动转换成工单的判断处理：读取生效的MRP配置
            r=mrp_rules.retrieve({"include_items":True})
            self.response_info(r)
            if len(r.json())==0:
                Logger.info("未有MRP配置")
                pytest.fail()
            isAutoChangeProductionPlan=0
            for mrpRule in r.json():
                if "global" in mrpRule:
                    for item in mrpRule["items"]:
                        if item["demand_type"]=="MAKE_PLAN_AUTO_RELEASE":
                            isAutoChangeProductionPlan=item["kept"]
            if isAutoChangeProductionPlan==1:#自动转MO。此时应该是已经转了MO，不需要查需求，直接进行校验
                for product_info in self.case_data["数据"][3]["products"]:
                    #需要校验的数据有：销售订单名称，参考单据类型，客户名称，订单交期，合同编号，产品编号，生产数量，计划完成日期，工单bom关系，物料需求明细
                    

                    #查找对应工单
                    data={}
                    data["material_ids"]=[product_info["material_id"]]
                    #data["so_id"]=so_id
                    #下发搜索信息
                    data["kitting_status"]=0
                    data["shipped_status"]=0
                    data["finish_status"]=0
                    data["need_update_bom"]=False
                    data["production_plan_status"]=[0,10]
                    data["size"]=20
                    data["page"]=0
                    
                    r=work_order.retrieve(data)
                    self.response_info(r)
                    if r.json()["total_elements"]==0:
                        Logger.info("未查到物料编号为"+product_info["material_no"]+"的工单")
                        pytest.fail()
                    work_order_id=r.json()["content"][0]['id']
                    r=work_order.get_detail({"production_plan_id":work_order_id})
                    self.response_info(r)

                    #关于基本校验信息的包装
                    check_info={}
                    check_info["参考单据类型"]=r.json()["reference_document_type_desc"]
                    check_info["生产数量"]=r.json()["total_quantity"]
                    check_info["计划完成日期"]=r.json()["plan_stock_date"]

                    #关于BOM校验信息的包装
                    check_bom=material.get_bom_from_wo_detail(r.json())#通过该方法获取查询到的工单详情的bom信息，返回一个如[{'A': [{'B': 1}, {'C': 1}]}]的列表去表示bom

                    #关于物料明细校验信息的包装：直接由测试数据得出应有的物料明细即可,给出内容为：物料编号，工单需求，单台需求
                    check_material_info=[]
                    if len(r.json()["bom_qty_list"])!=0:
                        for bom_qty in r.json()["bom_qty_list"]:
                            tmp={}
                            tmp["物料编号"]=bom_qty["material_no"]
                            tmp["工单需求"]=bom_qty["required_qty"]
                            tmp["单台需求"]=bom_qty["required_unit_qty"]
                            check_material_info.append(tmp)
                    material_no=product_info["material_no"]
                    bom_compare=0
                    is_check_data_bom=0
                    for check_data in self.case_data['期望校验数据']:
                        if check_data["交付物"]==material_no:
                            self.compare_result_detail(check_info,check_data)#校验基本信息
                            BOM_info=material.get_bom_from_casedata(check_data["BOM"])#获取物料BOM的关系
                            bom_compare=self.is_list_same(check_bom,BOM_info)#比较物料BOM是否一致
                            if bom_compare==0:
                                Logger.info("工单BOM与期望数据不一致")
                                pytest.fail()
                            #比较物料明细信息是否一致
                            if len(check_data["物料明细"])==0 and len(check_material_info)==0:#没有给出期望的物料明细信息以及工单详情没有物料明细信息
                                pass
                            elif len(check_data["物料明细"])==0:
                                Logger.info("期望数据未给出产出物"+material_no+"的物料明细期望校验值")
                                pytest.fail()
                            elif len(check_material_info)==0:
                                Logger.info(material_no+"给出物料明细期望校验值，但工单内未有物料明细值")
                                pytest.fail()
                            elif self.is_list_same(check_material_info,check_data["物料明细"])==0:
                                Logger.info(material_no+"的物料明细期望校验值与实际不一致")
                                pytest.fail()
                            is_check_data_bom=1
                    if is_check_data_bom==0:
                        Logger.info("期望数据未给出产出物"+material_no+"的期望校验值")
                        pytest.fail()

                    matids=[]
                    matids.append(product_info["material_id"])
                    demand_info={
                    "demand_type":"PENDING_PLAN",
                    "procurement_type":"MO",
                    "material_ids": matids,
                    "only_exceptions": False,
                    "size":20,
                    "page":0
                    }
                    #当转换工单后，原先的请求应该已经没了
                    r=demand_supply.get_demand_supply(demand_info)
                    self.response_info(r)
                    assert r.json()["total_elements"] == 0

            else:#未勾选自动转工单，需要手动转换
                #查询生成的生产需求,获取生产需求id和计划开始计划完成时间
                for product_info in self.case_data["数据"][3]["products"]:
                    matids=[]
                    matids.append(product_info["material_id"])
                    demand_info={
                        "demand_type":"PENDING_PLAN",
                        "procurement_type":"MO",
                        "material_ids": matids,
                        "only_exceptions": False,
                        "size":20,
                        "page":0
                    }
                    r=demand_supply.get_demand_supply(demand_info)
                    self.response_info(r)
                    assert r.json()["total_elements"]>0
                    demand_id=r.json()['content'][0]['id']#此处先直接能准确查出一个结果来算
                    ds_date=r.json()['content'][0]['ds_date']
                    prod_date=r.json()['content'][0]['prod_date']
                    material_id=product_info["material_id"]
                    material_name=product_info["material_name"]
                    material_no=product_info["material_no"]
                    #获取区域，id和名字
                    r=area.retrieve_areas({"material_no":material_no})
                    self.response_info(r)
                    area_id=r.json()[0]["id"]
                    area_name=r.json()[0]["area_name"]
                    #获取工艺路线，id和名字
                    r=route.retrieve_routes({"material_no":material_no})
                    self.response_info(r)
                    route_id=r.json()[0]["id"]
                    route_name=r.json()[0]["route_name"]
                    #获取主数据物料BOM树
                    r=work_order.retrieve_bomm_tree([material_id])
                    self.response_info(r)
                    if "children" in r.json()[0]:#存在所添加的主数据没有子物料的情况，此处进行了处理。
                        children_list=r.json()[0]["children"]
                        material.recursiom_mixed_bom_data(children_list,1,0,None,[])
                        kitting_status=0
                    else:
                        children_list=[]
                        kitting_status=10
                    Logger.info(children_list)

                
                    #获取订单信息，交付物数量，id，no，名字，交期
                    r=sales_order.get_detail({'so_id': so_id})
                    batch_delivery_date=r.json()["batch_delivery_date"]
                    so_no=r.json()["batch_no"]
                    so_name=r.json()["batch_name"]
                    #不同的交付物信息不一样，需要处理
                    document_item_id=""
                    total_quantity=0
                    for prod in r.json()["products"]:
                        if prod["material_id"]==material_id:
                            document_item_id=prod["id"]+document_item_id
                            total_quantity=prod["quantity"]+total_quantity
                    #包装请求
                    data={}
                    data["reference_document_type"]="HAND_TIED_SO"
                    data["contract_name"]=contract_name
                    data["contract_no"]=contract_no
                    data["account_name"]=account_name
                    data["material_id"]=material_id
                    data["material_name"]=material_name
                    data["material"]={
                        "id":material_id,
                        "material_name":material_name,
                        "material_no":material_no
                    }
                    data["mo_type"]="PRODUCTION_ORDER"
                    data["pe_route_id"]=route_id
                    data["pe_area_id"]=area_id
                    data["plan_prod_date"]=prod_date
                    data["plan_stock_date"]=ds_date
                    data["document_date"]=batch_delivery_date
                    data["document_type"]="SO"
                    data["document_id"]=so_id
                    data["document_no"]=so_no
                    data["document_name"]=so_name
                    data["document_item_id"]=document_item_id
                    data["pe_route_name"]=route_name
                    data["pe_area_name"]=area_name
                    data["completed_quantity"]=0
                    data["kitting_status"]=kitting_status
                    data["kitting_items"]=children_list
                    data["total_quantity"]=total_quantity
                    data["demand_id"]=demand_id
                    r=work_order.create(data)
                    self.response_info(r)
                    #需要校验的数据有：销售订单名称，参考单据类型，客户名称，订单交期，合同编号，产品编号，生产数量，计划完成日期，工单bom关系，物料需求明细
                    work_order_id=r.json()["id"]
                    r=work_order.get_detail({"production_plan_id":work_order_id})
                    self.response_info(r)

                    #关于基本校验信息的包装
                    check_info={}
                    check_info["销售订单名称"]=r.json()["document_name"]
                    check_info["参考单据类型"]=r.json()["reference_document_type_desc"]
                    check_info["客户名称"]=r.json()["account_name"]
                    check_info["订单交期"]=r.json()["document_date"]
                    check_info["合同编号"]=r.json()["contract_no"]
                    check_info["生产数量"]=r.json()["total_quantity"]
                    check_info["计划完成日期"]=r.json()["plan_stock_date"]

                    #关于BOM校验信息的包装
                    check_bom=material.get_bom_from_wo_detail(r.json())#通过该方法获取查询到的工单详情的bom信息，返回一个如[{'A': [{'B': 1}, {'C': 1}]}]的列表去表示bom

                    #关于物料明细校验信息的包装：直接由测试数据得出应有的物料明细即可,给出内容为：物料编号，工单需求，单台需求
                    check_material_info=[]
                    if len(r.json()["bom_qty_list"])!=0:
                        for bom_qty in r.json()["bom_qty_list"]:
                            tmp={}
                            tmp["物料编号"]=bom_qty["material_no"]
                            tmp["工单需求"]=bom_qty["required_qty"]
                            tmp["单台需求"]=bom_qty["required_unit_qty"]
                            check_material_info.append(tmp)

                    bom_compare=0
                    is_check_data_bom=0
                    for check_data in self.case_data['期望校验数据']:
                        if check_data["交付物"]==material_no:
                            self.compare_result_detail(check_info,check_data)#校验基本信息
                            BOM_info=material.get_bom_from_casedata(check_data["BOM"])#获取物料BOM的关系
                            bom_compare=self.is_list_same(check_bom,BOM_info)#比较物料BOM是否一致
                            if bom_compare==0:
                                Logger.info("工单BOM与期望数据不一致")
                                pytest.fail()
                            #比较物料明细信息是否一致
                            if len(check_data["物料明细"])==0 and len(check_material_info)==0:#没有给出期望的物料明细信息以及工单详情没有物料明细信息
                                pass
                            elif len(check_data["物料明细"])==0:
                                Logger.info("期望数据未给出产出物"+material_no+"的物料明细期望校验值")
                                pytest.fail()
                            elif len(check_material_info)==0:
                                Logger.info(material_no+"给出物料明细期望校验值，但工单内未有物料明细值")
                                pytest.fail()
                            elif self.is_list_same(check_material_info,check_data["物料明细"])==0:
                                Logger.info(material_no+"的物料明细期望校验值与实际不一致")
                                pytest.fail()
                            is_check_data_bom=1
                    if is_check_data_bom==0:
                        Logger.info("期望数据未给出产出物"+material_no+"的期望校验值")
                        pytest.fail()

                    #当转换工单后，原先的请求应该已经没了
                    r=demand_supply.get_demand_supply(demand_info)
                    self.response_info(r)
                    assert r.json()["total_elements"] == 0
                

    class update_bom_of_workorder(Abstract_case): 
        def __init__(self, text=None) -> None:
            super().__init__('计划管理/工单管理/单个更新工单BOM', text)
        def run(self):
            #前置条件：已创建可用工单。测试数据应给出被修改的物料id，对应的销售订单，新的BOM.Bom应根据已有的物料去修改，此处在测试数据的供给上需要注意！
            #修改指定物料的BOM
            BOM_info=material.get_bom_from_casedata(self.case_data['数据'][2]["BOM"])#获取物料BOM的关系
            #根据BOM去维护系统中物料的BOM的关系
            if len(BOM_info)!=0:
                for bom_info in BOM_info:
                    material.change_bom(bom_info)
            #在主数据修改完BOM后，需要在对应工单去转换
            #根据测试数据（销售订单名称+物料名称）进行查询工单，返回的数据与校验数据（销售订单名称+物料id）进行对比
            #查询销售订单名称查询，获取销售订单的信息
            r=sales_order.retrieve(self.case_data['数据'][1])
            self.response_info(r)
            if r.json()["total_elements"] == 0:
                Logger.error('查询的订单未创建')
                pytest.fail()
            #此处存在查询会返回多个值的问题，获取返回值中与查询数据完全一致的数据
            so_info=self.query_directly(r.json()["content"],"batch_name",self.case_data['数据'][1]["销售订单名称"])
            data={}
            data["so_id"]=so_info["id"]
            #查询物料信息
            r=material.retrieve(self.case_data["数据"][0])
            self.response_info(r)
            if r.json()["total_elements"] == 0:
                Logger.error('未在系统准备必要的物料')
                pytest.fail()
            #此处存在查询会返回多个值的问题，获取返回值中与查询数据完全一致的数据
            material_info=self.query_directly(r.json()["content"],"material_name",self.case_data['数据'][0]["物料名称"])
            data["material_ids"]=[material_info["id"]]
            #下发搜索信息
            data["kitting_status"]=0
            data["shipped_status"]=0
            data["finish_status"]=0
            data["need_update_bom"]=False
            data["production_plan_status"]=[0,10]
            data["size"]=20
            data["page"]=0
            r=work_order.retrieve(data)
            self.response_info(r)
            work_order_id=r.json()["content"][0]['id']
            r=work_order.mo_bom({"production_plan_id":work_order_id})
            self.response_info(r)
            #校验
            #此处校验实际上依赖于测试数据
            r=work_order.get_detail({"production_plan_id":work_order_id})
            self.response_info(r)
            #关于BOM校验信息的包装
            check_bom=material.get_bom_from_wo_detail(r.json())#通过该方法获取查询到的工单详情的bom信息，返回一个如[{'A': [{'B': 1}, {'C': 1}]}]的列表去表示bom
            BOM_info=material.get_bom_from_casedata(self.case_data["期望校验数据"][0]["BOM"])#获取物料BOM的关系
            bom_compare=self.is_list_same(check_bom,BOM_info)#比较物料BOM是否一致
            if bom_compare==0:
                Logger.info("工单BOM与期望数据不一致")
                pytest.fail()


    class release_workorder(Abstract_case): 
        def __init__(self, text=None) -> None:
            super().__init__('计划管理/工单管理/释放工单', text)
        def run(self):
             #根据测试数据（销售订单名称+物料名称）进行查询工单，返回的数据与校验数据（销售订单名称+物料id）进行对比
            #查询销售订单名称查询，获取销售订单的信息
            r=sales_order.retrieve(self.case_data['数据'][1])
            self.response_info(r)
            if r.json()["total_elements"] == 0:
                Logger.error('查询的订单未创建')
                pytest.fail()
            #此处存在查询会返回多个值的问题，获取返回值中与查询数据完全一致的数据
            so_info=self.query_directly(r.json()["content"],"batch_name",self.case_data['数据'][1]["销售订单名称"])
            data={}
            data["so_id"]=so_info["id"]
            #查询物料信息
            r=material.retrieve(self.case_data["数据"][0])
            self.response_info(r)
            if r.json()["total_elements"] == 0:
                Logger.error('未在系统准备必要的物料')
                pytest.fail()
            #此处存在查询会返回多个值的问题，获取返回值中与查询数据完全一致的数据
            material_info=self.query_directly(r.json()["content"],"material_name",self.case_data['数据'][0]["物料名称"])
            data["material_ids"]=[material_info["id"]]
            #下发搜索信息
            data["kitting_status"]=0
            data["shipped_status"]=0
            data["finish_status"]=0
            data["need_update_bom"]=False
            data["production_plan_status"]=[0,10]
            data["size"]=20
            data["page"]=0
            r=work_order.retrieve(data)
            self.response_info(r)
            if "pe_area_id" in r.json()["content"][0] and "pe_route_id" in r.json()["content"][0]:
                pass
            else:
                Logger.error('该工单无法释放，区域、工艺路线没有设置')
                pytest.fail()
            work_order_id=r.json()["content"][0]['id']
            r=work_order.release({"production_plan_id":work_order_id})
            self.response_info(r)
            #校验
            #释放后的工单详情应该多2个字段：release_date和status改为10
            r=work_order.get_detail({"production_plan_id":work_order_id})
            self.response_info(r)
            assert r.json()["status"]==10
            today=time.strftime("%Y-%m-%d", time.localtime())
            Logger.info(today)
            assert r.json()["release_date"]==today

    class finish_workorder(Abstract_case): 
        def __init__(self, text=None) -> None:
            super().__init__('计划管理/工单管理/工单完工', text)
        def run(self):
            #在点击完工按钮时，有一个wm的查询link对应sn的接口，目前暂不知该接口的意义，此处先留空
            #批量获取工单详情
            
            #根据测试数据（销售订单名称+物料名称）进行查询工单，返回的数据与校验数据（销售订单名称+物料id）进行对比
            #查询销售订单名称查询，获取销售订单的信息
            r=sales_order.retrieve(self.case_data['数据'][1])
            self.response_info(r)
            if r.json()["total_elements"] == 0:
                Logger.error('查询的订单未创建')
                pytest.fail()
            #此处存在查询会返回多个值的问题，获取返回值中与查询数据完全一致的数据
            so_info=self.query_directly(r.json()["content"],"batch_name",self.case_data['数据'][1]["销售订单名称"])
            data={}
            data["so_id"]=so_info["id"]
            #查询物料信息
            r=material.retrieve(self.case_data["数据"][0])
            self.response_info(r)
            if r.json()["total_elements"] == 0:
                Logger.error('未在系统准备必要的物料')
                pytest.fail()
            #此处存在查询会返回多个值的问题，获取返回值中与查询数据完全一致的数据
            material_info=self.query_directly(r.json()["content"],"material_name",self.case_data['数据'][0]["物料名称"])
            data["material_ids"]=[material_info["id"]]
            #下发搜索信息
            data["kitting_status"]=0
            data["shipped_status"]=0
            data["finish_status"]=0
            data["need_update_bom"]=False
            data["production_plan_status"]=[0,10]
            data["size"]=20
            data["page"]=0
            r=work_order.retrieve(data)
            self.response_info(r)
            work_order_info=self.query_directly(r.json()['content'],'document_name',self.case_data['数据'][1]["销售订单名称"])
            if work_order_info["status"]==0:#若工单未被释放，则无法完工
                Logger.error('该工单仍未释放！')
                pytest.fail()
            work_order_id=work_order_info["id"]
            #根据工单id批量获取工单详情
            wo_info=[]
            wo_info.append(work_order_id)
            r=work_order.get_detail_bulk_by_id(wo_info)
            self.response_info(r)

            #保存生产进度
            progress=[]
            progress_info={
                "id":work_order_id,
                "finish_mode":"FINISH_TOTALLY"
            }
            progress.append(progress_info)
            r=work_order.save_progress(progress)
            self.response_info(r)

            #判断是否该工单可以被完工
            r=work_order.finish_check(wo_info)
            self.response_info(r)
            if len(r.json()["exception_production_plans"])!=0:
                Logger.error('该工单无法被完工')
                pytest.fail()

            #执行完工
            self.case_data['数据'][2]["production_plans"]=[]
            del progress_info["finish_mode"]
            progress_info["exception_handles"]=None
            self.case_data['数据'][2]["production_plans"].append(progress_info)
            r=mo.finish_process(self.case_data['数据'][2])
            self.response_info(r)
            #校验
            #完工后的工单多一个actual_stock_date，finish_status为1
            r=work_order.get_detail({"production_plan_id":work_order_id})
            self.response_info(r)
            check_data={}
            check_data["完工时间"]=r.json()["actual_stock_date"]
            check_data["finish_status"]=r.json()["finish_status"]
            self.case_data['期望校验数据'][0]["finish_status"]=1
            self.compare_result_detail(check_data,self.case_data['期望校验数据'][0])

    class create_MO(Abstract_case): 
        def __init__(self, text=None) -> None:
            super().__init__('计划管理/工单入库申请单/工单入库申请单新建', text)
        def run(self):
            #实际步骤：选择已经完工的工单，填入必填信息（申请入库数量）提交，校验工单入库申请单信息是否正确
            #special：多工单，申请入库数量小于等于完工数量-已入库数量-已开申请单但未完成数量
            #前置条件：所要入库的工单已经创建并完工
            #根据测试数据给出的工单，先校验是否完工，未完工则报错
            mo_warehouse_application_item_create_list=[]
            for workorder in self.case_data['数据']:
                data={}
                if "销售订单名称" in workorder:#存在没有手绑销售订单工单的情况
                    r=sales_order.retrieve({"销售订单名称":workorder["销售订单名称"]})
                    self.response_info(r)
                    if r.json()["total_elements"] == 0:
                        Logger.error('查询的订单未创建')
                        pytest.fail()
                    #此处存在查询会返回多个值的问题，获取返回值中与查询数据完全一致的数据
                    so_info=self.query_directly(r.json()["content"],"batch_name",workorder["销售订单名称"])
                    data["so_id"]=so_info["id"]
                else:
                    data["so_id"]=None
                #查询物料信息
                r=material.retrieve({"物料名称":workorder["物料名称"]})
                self.response_info(r)
                if r.json()["total_elements"] == 0:
                    Logger.error('未在系统准备必要的物料')
                    pytest.fail()
                #此处存在查询会返回多个值的问题，获取返回值中与查询数据完全一致的数据
                material_info=self.query_directly(r.json()["content"],"material_name",workorder["物料名称"])
                data["material_ids"]=[material_info["id"]]
                #下发搜索信息
                data["kitting_status"]=0
                data["shipped_status"]=0
                data["finish_status"]=1
                data["need_update_bom"]=False
                data["production_plan_status"]=[0,10]
                data["size"]=20
                data["page"]=0
                r=work_order.retrieve(data)
                self.response_info(r)
                if r.json()["total_elements"]==0:
                    Logger.error('该工单仍未完工！')
                    pytest.fail()
                work_order_info=self.query_directly(r.json()['content'],'material_id',material_info["id"])
                #查看该工单的可入库数量是否大于等于申请入库数量，若小于，则报错
                work_order_id=work_order_info["id"]
                
                i=0
                for check in self.case_data['期望校验数据']:
                    if check == workorder:#此处一开始的期望校验数据应与给出的查询工单数据以及申请入库数量是一致的
                        check["工单号"]=work_order_info["sn"]
                        i=1
                        break
                if i==0:
                    Logger.error('期望校验数据与测试数据未一一对应')
                    pytest.fail()

                r=mo_warehouse_application.bulk({"mo_ids":[work_order_id]})
                self.response_info(r)
                if r.json()[0]["pending_delivery_quantity"]<workorder["申请入库数量"]:
                    Logger.error('申请入库数量大于可申请入库数量')
                    pytest.fail()
                mo_warehouse_application_item_create={}
                mo_warehouse_application_item_create["mo_id"]=work_order_id
                mo_warehouse_application_item_create["quantity"]=workorder["申请入库数量"]
                mo_warehouse_application_item_create["priority"]=None
                mo_warehouse_application_item_create["self_make_inbound_storage_location_id"]=None
                mo_warehouse_application_item_create["remark"]=None
                mo_warehouse_application_item_create_list.append(mo_warehouse_application_item_create)
            r=mo_warehouse_application.create({"mo_warehouse_application_item_create_list":mo_warehouse_application_item_create_list})
            self.response_info(r,"创建工单入库申请单")
            #校验：参考数据为物料编号，工单号，销售订单号，考虑到有些工单未绑销售订单，在给出期望校验数据时，可以给出物料编号，工单号，销售订单号（选填），申请入库数量
            mo_warehouse_application_id=r.json()["id"]
            if "pp" in global_data.data:
                global_data.data["pp"]["mo_warehouse_application_id"]=mo_warehouse_application_id
            else:
                global_data.data["pp"]={}
                global_data.data["pp"]["mo_warehouse_application_id"]=mo_warehouse_application_id
            r=mo_warehouse_application.get_detail({"mo_warehouse_application_id":mo_warehouse_application_id})
            self.response_info(r,"查询工单入库申请单详情")
            
            check_info_list=[]
            for mo_warehouse_application_item in r.json()["mo_warehouse_application_item_list"]:#多工单入库申请行
                check_info={}
                check_info["物料名称"]=mo_warehouse_application_item["material_name"]
                check_info["工单号"]=mo_warehouse_application_item["mo_no"]
                check_info["申请入库数量"]=mo_warehouse_application_item["target_quantity"]
                if "so_name" in mo_warehouse_application_item:
                    check_info["销售订单名称"]=mo_warehouse_application_item["so_name"]
                check_info_list.append(check_info)
            if self.is_list_same(check_info_list,self.case_data["期望校验数据"]) == 0:
                Logger.error('工单入库申请单数据与测试数据不符')
                pytest.fail()


    class retrieve_MO(Abstract_case): 
        def __init__(self, text=None) -> None:
            super().__init__('计划管理/工单入库申请单/工单入库申请单查询', text)
        def run(self):
            #实际步骤：给出查询条件，查询，校验查询出的结果是否正确
            #给出的测试数据应该是：销售订单和物料号，确定到工单号，然后根据工单号查询，查询结果可能不止一个，并且一个工单入库申请单内的内容可能不止一个，校验方式是，查看每个申请单是否包含了该工单即可
            #该用例限制给出销售订单和物料信息
            data={}
            r=sales_order.retrieve({"销售订单名称":self.case_data["数据"][0]["销售订单名称"]})
            self.response_info(r,"查询销售订单")
            if r.json()["total_elements"] == 0:
                Logger.error('查询的订单未创建')
                pytest.fail()
            #此处存在查询会返回多个值的问题，获取返回值中与查询数据完全一致的数据
            so_info=self.query_directly(r.json()["content"],"batch_name",self.case_data["数据"][0]["销售订单名称"])
            data["so_id"]=so_info["id"]
            r=material.retrieve({"物料名称":self.case_data["数据"][0]["物料名称"]})
            self.response_info(r,"查询物流")
            if r.json()["total_elements"] == 0:
                Logger.error('未在系统准备必要的物料')
                pytest.fail()
            #此处存在查询会返回多个值的问题，获取返回值中与查询数据完全一致的数据
            material_info=self.query_directly(r.json()["content"],"material_name",self.case_data["数据"][0]["物料名称"])
            data["material_ids"]=[material_info["id"]]
            #下发搜索信息
            data["kitting_status"]=0
            data["shipped_status"]=0
            data["finish_status"]=1
            data["need_update_bom"]=False
            data["production_plan_status"]=[0,10]
            data["size"]=20
            data["page"]=0
            r=work_order.retrieve(data)
            self.response_info(r,"查询工单")
            if r.json()["total_elements"]==0:
                Logger.error('该工单仍未完工！')
                pytest.fail()
            work_order_info=self.query_directly(r.json()['content'],'material_id',material_info["id"])
            work_order_no=work_order_info["sn"]
            r=mo_warehouse_application.retrieve({"pp_no":work_order_no})
            self.response_info(r,"查询工单入库申请单")
            #校验：实际上按照工单号筛选出来的结果在申请单列表无法直观看到，需要查看申请单详情。由于工单查询是限制了必须是带了销售订单和物料的情况，需要
            if r.json()["total_elements"] == 0:
                Logger.error("未查到对应的工单入库申请单")
                pytest.fail()
            self.case_data["期望校验数据"][0]["工单号"]=work_order_no
            total_elements=r.json()["total_elements"]#查询出的结果的条目数
            is_application_cover_work_order=0
            for application in r.json()["content"]:
                mo_warehouse_application_id=application["id"]
                resp=mo_warehouse_application.get_detail({"mo_warehouse_application_id":mo_warehouse_application_id})
                self.response_info(resp,"查询工单入库申请单详情")
                is_cover_work_order=0
                for mo_warehouse_application_item in resp.json()["mo_warehouse_application_item_list"]:
                    check_data={}
                    check_data["销售订单名称"]=mo_warehouse_application_item["so_name"]
                    check_data["物料名称"]=mo_warehouse_application_item["material_name"]
                    check_data["工单号"]=mo_warehouse_application_item["mo_no"]
                    if check_data==self.case_data["期望校验数据"][0]:
                        is_cover_work_order=1
                        break
                if is_cover_work_order==0:
                    Logger.error("查询结果中入库申请单"+resp.json()["order_no"]+"未包含校验数据所查询的工单")
                    pytest.fail()
                else:
                    is_application_cover_work_order=is_application_cover_work_order+1
            assert is_application_cover_work_order == total_elements
    
    class detail_MO(Abstract_case): 
        def __init__(self, text=None) -> None:
            super().__init__('计划管理/工单入库申请单/工单入库申请单详情', text)
        def run(self):
            #实际步骤：给出查询条件，查询，获取工单入库申请单id，获取详情，校验数据是否正确
            #在测试数据中应该是无法给出工单号、工单入库申请单号的，所以，在只能给出销售订单名称和物料的情况下，可以获得工单，查看到的该工单对应的工单入库申请单，只要该工单入库申请单里包含了该工单即可算是正确
            data={}
            r=sales_order.retrieve({"销售订单名称":self.case_data["数据"][0]["销售订单名称"]})
            self.response_info(r,"查询销售订单")
            if r.json()["total_elements"] == 0:
                Logger.error('查询的订单未创建')
                pytest.fail()
            #此处存在查询会返回多个值的问题，获取返回值中与查询数据完全一致的数据
            so_info=self.query_directly(r.json()["content"],"batch_name",self.case_data["数据"][0]["销售订单名称"])
            data["so_id"]=so_info["id"]
            r=material.retrieve({"物料名称":self.case_data["数据"][0]["物料名称"]})
            self.response_info(r,"查询物流")
            if r.json()["total_elements"] == 0:
                Logger.error('未在系统准备必要的物料')
                pytest.fail()
            #此处存在查询会返回多个值的问题，获取返回值中与查询数据完全一致的数据
            material_info=self.query_directly(r.json()["content"],"material_name",self.case_data["数据"][0]["物料名称"])
            data["material_ids"]=[material_info["id"]]
            #下发搜索信息
            data["kitting_status"]=0
            data["shipped_status"]=0
            data["finish_status"]=1
            data["need_update_bom"]=False
            data["production_plan_status"]=[0,10]
            data["size"]=20
            data["page"]=0
            r=work_order.retrieve(data)
            self.response_info(r,"查询工单")
            if r.json()["total_elements"]==0:
                Logger.error('该工单仍未完工！')
                pytest.fail()
            work_order_info=self.query_directly(r.json()['content'],'material_id',material_info["id"])
            work_order_no=work_order_info["sn"]
            r=mo_warehouse_application.retrieve({"pp_no":work_order_no})
            self.response_info(r,"查询工单入库申请单")
            #校验：实际上按照工单号筛选出来的结果在申请单列表无法直观看到，需要查看申请单详情。由于工单查询是限制了必须是带了销售订单和物料的情况，需要
            if r.json()["total_elements"] == 0:
                Logger.error("未查到对应的工单入库申请单")
                pytest.fail()
            application=r.json()["content"][0]
            mo_warehouse_application_id=application["id"]
            resp=mo_warehouse_application.get_detail({"mo_warehouse_application_id":mo_warehouse_application_id})
            self.response_info(resp,"查询工单入库申请单详情")
            is_cover_work_order=0
            for mo_warehouse_application_item in resp.json()["mo_warehouse_application_item_list"]:
                check_data={}
                check_data["销售订单名称"]=mo_warehouse_application_item["so_name"]
                check_data["物料名称"]=mo_warehouse_application_item["material_name"]
                if check_data==self.case_data["期望校验数据"][0]:
                    is_cover_work_order=1
                    break
            if is_cover_work_order==0:
                Logger.error("查询结果中入库申请单"+resp.json()["order_no"]+"未包含校验数据所查询的工单")
                pytest.fail()
            
    class remove_MO(Abstract_case): 
        def __init__(self, text=None) -> None:
            super().__init__('计划管理/工单入库申请单/工单入库申请单废弃', text)
        def run(self):
            #实际步骤：给出查询条件，查询，获取工单入库申请单id，废弃申请单，校验
            #此处需要使用公用数据的新建的工单入库申请单数据，定位到对应的工单入库申请单
            if "pp" in global_data.data and "mo_warehouse_application_id" in global_data.data["pp"]:
                mo_warehouse_application_id=global_data.data["pp"]["mo_warehouse_application_id"]
            else:
                Logger.error("未有创建的工单入库申请单可进行废弃")
                pytest.fail()
            r=mo_warehouse_application.abandon({"mo_warehouse_application_id":mo_warehouse_application_id})
            self.response_info(r,"废弃工单入库申请单")
            r=mo_warehouse_application.get_detail({"mo_warehouse_application_id":mo_warehouse_application_id})
            self.response_info(r,"查询工单入库申请单详情")
            #校验
            #需要校验的数据为 status：O->A 
            check_data={"申请单状态":r.json()["status"]}
            assert check_data == self.case_data["期望校验数据"][0]

    class MRP_test(Abstract_case): 
        def __init__(self, text=None) -> None:
            super().__init__('计划管理/MRP/MRP结果校验场景一', text)
        
        def get_source_type(self,source_type,source_sub_type):
            if source_type== "1" and source_sub_type=="SALES_ORDER":
                return "销售订单"
            if source_type=="1" and source_sub_type=="RELATE_REQUEST":
                return "相关需求"
            if source_type=="2" and source_sub_type=="PENDING_PLAN":
                return "预排计划"
        
        def parent_ds_no_In_list(self,parent_ds_nos,list):
            #查看某个字典是否存在于列表中
            for l in list:
                if parent_ds_nos == l["ds_no"]:
                    return 1
            return 0
        
        def parent_node_In_result(self,parent_ds_nos,result_list,node):#该方法判断是否父节点在result_list里
            judge=0
            for result in result_list:
                if parent_ds_nos==result["ds_no"]:
                    if "产生" in result:
                        result["产生"].append(node)
                    else:
                        result["产生"]=[node]
                    judge=judge+1
                    break
                else:
                    if "产生" in result:
                        judge=self.parent_node_In_result(parent_ds_nos,result["产生"],node)
                    if judge==1:
                        break
            return judge


        def mix_node_data(self,node_list:list,node_list_tmp:list,result_list):#该方法用于包装node_list中的数据
            if len(node_list_tmp)!=0:
                node=node_list_tmp[0]
                if "parent_ds_nos" not in node:#没有父节点
                    result_list.append(node)
                    node_list_tmp.pop(0)
                else:
                    for parent_ds_nos in node["parent_ds_nos"]:
                        if self.parent_ds_no_In_list(parent_ds_nos,node_list)==0:#父节点在列表中找不到
                            result_list.append(node)
                        else:
                            if self.parent_node_In_result(parent_ds_nos,result_list,node)==1:#父节点已经在result列表中
                                pass
                            else:
                                for n in node_list_tmp:
                                    if n["ds_nos"]==parent_ds_nos:
                                        n["产生"]=[node]
                                        break
                    node_list_tmp.pop(0)
                self.mix_node_data(node_list,node_list_tmp,result_list)
            else:
                pass                

        def hanlde_data(self,result_list):#该方法用于处理result_list
            for r in result_list:
                del r["ds_no"]
                if "parent_ds_nos" in r:
                    del r["parent_ds_nos"]
                if "source_number" in r and r["类型"]=="销售订单":
                    resp=sales_order.retrieve({"so_query":r["source_number"]})
                    r["单据名称"]=resp.json()["content"][0]["batch_name"]
                    del r["source_number"]
                if r["类型"]=="预排计划":
                    if "产生" in r:
                        for rr in r["产生"]:
                            if rr["类型"] == "相关需求":
                                rr["类型"]="相关需求（预排）"
                if "产生" in r:
                    self.hanlde_data(r["产生"])


        def run(self):
            #实现步骤：创建物料，维护bom，创建销售订单并下发，跑全量MRP，查询各个物料的MRP结果是否正确
            #根据已创建的销售合同，创建销售订单（前置条件，需满足），获取销售订单id
            ordered_id = account.get_field_by_name(self.case_data['数据'][0],"id")
            r = contract.retrieve(self.case_data['数据'][1])
            self.response_info(r)
            if r.json()["total_elements"] == 0:
                Logger.error('未在系统准备必要的销售合同')
                pytest.fail()
            contract_id=r.json()["content"][0]["id"]
            customer_id=customer.get_field_by_name(self.case_data["数据"][2],"id")           
            self.case_data["数据"][3]["orderedBy"] = ordered_id
            self.case_data["数据"][3]["contractId"] = contract_id
            self.case_data["数据"][3]["account_id"] = customer_id

            #关于物料的维护：选择每次自动创建，并去更新其BOM。            
            #根据计量单位组，查询主计量单位和采购单位，这两个单位都由unit_name得来
            r=material.material_unit_groups_page(self.case_data["数据"][5])
            self.response_info(r)
            material_unit_groups_id=r.json()['content'][0]['id']
            r=material.material_units_retrieve_by_measure_unit_groups({"id":material_unit_groups_id})
            self.response_info(r)
            if len(r.json())==0:
                Logger.error('未在系统准备必要的主计量单位')
                pytest.fail()
            material_units_name=r.json()[0]["unit"]["unit_name"]
            material_list=[]
            #循环创建所需要的物料
            for caseData in self.case_data['数据']:
                if "物料名称" in caseData:
                    caseData["material_unit"]=material_units_name
                    caseData["material_purchase_unit"]=material_units_name
                    r=material.create(caseData)
                    self.response_info(r)
                    material_info={}
                    material_info["id"]=r.json()["id"]
                    material_info["material_no"]=r.json()["material_no"]
                    material_list.append(material_info)
            #修改物料的BOM
            BOM_info=material.get_bom_from_casedata(self.case_data['数据'][4]["BOM"])#获取物料BOM的关系
            #根据BOM去维护系统中物料的BOM的关系
            if len(BOM_info)!=0:
                for bom_info in BOM_info:
                    material.update_bom(bom_info)
            #物料相关信息的封装
            #搜索物料，获取其中的参数值
            self.case_data["数据"][3]["products"]=[]
            sequence=1
            if len(self.case_data["数据"][6]["交付物"])==0:
                Logger.error('未在测试数据准备交付物的数据')
                pytest.fail()
            for material_search in self.case_data["数据"][6]["交付物"]:#处理销售订单带多个交付物的情况
                r=material.retrieve({"material_no":material_search["物料编号"]})
                product={}
                product["material_group"] = r.json()["content"][0]["material_group"]
                product["material_desc"] = r.json()["content"][0]["material_desc"]
                product["material_name"] = r.json()["content"][0]["material_name"]
                product["material_no"] = r.json()["content"][0]["material_no"]
                product["material_id"] = r.json()["content"][0]["id"]
                product["measure_unit"] = r.json()["content"][0]["purchase_measure_unit"]
                product["unit_price"]=material_search["未税单价"]
                product["tax_rate"]=material_search["税率"]
                product["unit_tax_price"]=material_search["含税单价"]
                product["total_tax_price"]=material_search["价税合计"]
                product["quantity"]=material_search["数量"]
                product["required_delivery_date"]=material_search["交货日期"]
                product["price_mode"]="UNIT_PRICE"
                product["material_unit_name"]=r.json()["content"][0]["material_unit"]
                product["total_price"]=material_search["未税总额"]
                product["sequence"]= sequence
                sequence=sequence+1
                self.case_data["数据"][3]["products"].append(product)
            r = sales_order.create(self.case_data["数据"][3])
            self.response_info(r)
            assert (len(r.json()['new_id']) > 0)
            so_id=r.json()['new_id']
            #下发该订单
            #先默认系统内销售订单审批流是关闭的
            r=sales_order.create_approval({"task_config":None,"back_to_pr_flag":None,"so_id":so_id})
            self.response_info(r)
            #跑MRP,判断MRP计算是否成功，若成功继续，失败则报错
            r=task.create({"code":"MRP_COMPUTE"})
            self.response_info(r)
            task_id=r.json()["id"]
            r=task.retrieve({"size": 20,"page": 0})
            self.response_info(r)
            assert r.json()["total_elements"]!=0
            status="CREATED"
            while status == "CREATED":
                time.sleep(5)#考虑到可能查询时候任务未完成，设置5秒一查
                r=task.retrieve({"size": 20,"page": 0})
                for items in r.json()["content"]:
                    if items["id"]==task_id:
                        if items["status"]=="CREATED":
                            break
                        elif items["status"]=="EXECUTING":
                            break
                        elif items["status"]=="SUCCEEDED":
                            status="SUCCEEDED" 
                            break
                        else:
                            Logger.info("MRP计算失败")
                            pytest.fail()
            #此时MRP计算完成

            #查询各个物料的MRP结果，并进行校验
            material_ids=[]
            for ml in material_list:
                material_ids.append(ml["id"])
            data={}
            data["material_ids"]=material_ids
            data["page"]=0
            data["size"]=20
            r=demand_supply.get_demand_supply(data)
            self.response_info(r,"获取MRP列表")
            
            #根据返回结果，得到MRP的结果列表
            node_list=[]
            for cont in r.json()["content"]:
                node={}
                node["ds_no"]=cont["ds_no"]
                if "parent_ds_nos" in cont:
                    node["parent_ds_nos"] = cont["parent_ds_nos"]
                node["类型"]=self.get_source_type(cont["source_type"],cont["source_sub_type"])
                if "source_number" in cont:
                    node["source_number"] = cont["source_number"]
                node["物料编号"]=cont["material"]["material_no"]
                node["需求日期"]=cont["ds_date"]
                node["数量"]=cont["quantity"]
                if "prod_date" in cont:
                    node["建议开始日期"] = cont["prod_date"]
                node_list.append(node)
            
            result_list=[]
            node_list_tmp=[]
            for n in node_list:
                node_list_tmp.append(n)
            self.mix_node_data(node_list,node_list_tmp,result_list)
            #此时需要对result_list再做处理，类型需要完善（相关需求后加上字段），删除ds_no,parent_ds_no,source_number改为单据名称
            self.hanlde_data(result_list)
            Logger.info(result_list)
            self.is_list_same(result_list,self.case_data["期望校验数据"])

manager=Manager()